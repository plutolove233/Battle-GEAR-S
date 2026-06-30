#!/usr/bin/env python3
"""Generate a comprehensive Excel document describing hooks, atomic actions, and action card effects."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styling ──
header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
sub_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
category_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
category_font = Font(name="Microsoft YaHei", bold=True, size=11, color="2F5496")
body_font = Font(name="Microsoft YaHei", size=10)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def style_row(ws, row, max_col, font=None, fill=None, alignment=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        cell.border = thin_border
        if not alignment:
            cell.alignment = wrap_align


# ═══════════════════════════════════════════
# Sheet 1: 时点(Hook)一览表
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = "时点(Hook)一览"

headers1 = ["分类", "Hook常量名", "Hook值", "触发时机", "触发方式", "携带的Payload数据", "典型用途"]
for i, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, 1, len(headers1))

hooks_data = [
    # 游戏生命周期
    ("游戏生命周期", "HOOK_GAME_STARTED", "ON_GAME_STARTED", "游戏初始化完成时触发一次", "TurnService/GameSetupService", "{game_id, players}", "初始化效果、首回合准备"),
    ("游戏生命周期", "HOOK_GAME_OVER", "ON_GAME_OVER", "游戏结束条件达成时触发", "VictoryService", "{winner_id, reason}", "终局结算、结果统计"),

    # 轮次
    ("轮次", "HOOK_ROUND_START", "ON_ROUND_START", "新一轮开始时触发（双方各行动一次为一轮）", "RoundService", "{round_number}", "轮次级效果"),

    # 回合
    ("回合", "HOOK_TURN_START", "ON_TURN_START", "某方回合开始时触发。执行：抽牌、获得金币、恢复动力", "TurnService", "{player_id, turn_number}", "回合开始被动效果（如额外抽牌）"),
    ("回合", "HOOK_MAIN_PHASE_START", "ON_MAIN_PHASE_START", "主阶段开始时触发", "TurnService", "{player_id}", "主阶段自动触发的效果"),
    ("回合", "HOOK_OWNER_MAIN_PHASE", "ON_OWNER_MAIN_PHASE", "拥有者的主阶段触发（仅对效果拥有者生效）", "EffectEngine", "{player_id}", "主动效果发动窗口、主阶段可用能力"),
    ("回合", "HOOK_TURN_END", "ON_TURN_END", "回合结束时触发。执行：事件计时递减、弃牌、清除THIS_TURN效果", "TurnService", "{player_id, turn_number}", "回合结束效果（如未攻击回复威力）"),
    ("回合", "HOOK_OTHER_MECH_TURN_START", "ON_OTHER_MECH_TURN_START", "其他机甲回合开始时触发", "TurnService", "{other_mech_id, other_player_id}", "监视对手回合的效果"),

    # 卡牌
    ("卡牌", "HOOK_CARD_PLAYED", "ON_CARD_PLAYED", "任何卡牌被打出时触发", "CardPlayService", "{player_id, card_id, card_kind}", "通用打出效果"),
    ("卡牌", "HOOK_ACTION_CARD_PLAYED", "ON_ACTION_CARD_PLAYED", "行动牌被打出时触发", "CardPlayService", "{player_id, card_id, action_type}", "攻击牌打出时发起攻击声明"),
    ("卡牌", "HOOK_EQUIPMENT_SET", "ON_EQUIPMENT_SET", "装备牌被设置到区域时触发", "CardSetService", "{player_id, mech_id, card_id, slot_id}", "设置时触发效果（如抽牌）"),
    ("卡牌", "HOOK_EVENT_SET", "ON_EVENT_SET", "事件牌被设置到事件区域时触发", "CardSetService", "{player_id, mech_id, event_card_id, timer}", "事件牌设置效果"),
    ("卡牌", "HOOK_CARD_DISCARDED", "ON_CARD_DISCARDED", "卡牌被弃置时触发", "GameActions", "{card_id, owner_player_id, from_zone, reason}", "弃置反馈效果"),
    ("卡牌", "HOOK_CARD_DESTROYED", "ON_CARD_DESTROYED", "卡牌被破坏时触发（先触发此hook再弃置）", "GameActions", "{card_id, owner_player_id, reason}", "破坏结算效果"),

    # 攻击
    ("攻击", "HOOK_ATTACK_CARD_PLAYED", "ON_ATTACK_CARD_PLAYED", "攻击行动牌被打出时触发", "CardPlayService", "{player_id, card_id, weapon_id}", "发起攻击声明、不可无效等"),
    ("攻击", "HOOK_ATTACK_DECLARED", "ON_ATTACK_DECLARED", "攻击被声明后触发（选择武器和目标后）", "AttackService", "{attack_id, attacker_id, target_id, weapon_id}", "攻击声明时效果（如目标动力-2）"),
    ("攻击", "HOOK_ATTACK_RESPONSE_WINDOW", "ON_ATTACK_RESPONSE_WINDOW", "攻击响应窗口打开时触发，被攻击方可使用迎击牌", "AttackService", "{attack_id}", "回避/防御/识破等迎击效果"),
    ("攻击", "HOOK_ATTACK_MODIFIER_WINDOW", "ON_ATTACK_MODIFIER_WINDOW", "攻击修正窗口触发，可修改攻击威力和范围", "AttackService", "{attack_id, weapon_id, weapon_kind}", "威力+4/范围+1等攻击修正"),
    ("攻击", "HOOK_ATTACK_HIT", "ON_ATTACK_HIT", "攻击命中时触发", "AttackService", "{attack_id, target_id, damage}", "命中追加效果（额外损伤、弃牌等）"),
    ("攻击", "HOOK_ATTACK_MISS", "ON_ATTACK_MISS", "攻击未命中时触发", "AttackService", "{attack_id, target_id}", "未命中惩罚效果（如自损）"),
    ("攻击", "HOOK_ATTACK_RESOLVED", "ON_ATTACK_RESOLVED", "攻击结算完成后触发", "AttackService", "{attack_id, attacker_id, target_id}", "结算后效果（如反击、回复动力）"),
    ("攻击", "HOOK_MECH_TARGETED_BY_ATTACK", "ON_MECH_TARGETED_BY_ATTACK", "机甲被指定为攻击目标时触发（防守方）", "AttackService", "{target_id, attack_id}", "被攻击时效果（动力+2/护甲+5）"),
    ("攻击", "HOOK_MECH_HIT_BY_ATTACK", "ON_MECH_HIT_BY_ATTACK", "机甲被攻击命中时触发（防守方）", "AttackService", "{target_id, attack_id, damage}", "被命中效果（减伤、弃置装备减伤）"),
    ("攻击", "HOOK_REACTION_CARD_PLAYED", "ON_REACTION_CARD_PLAYED", "打出迎击牌时触发", "BattleState", "{player_id, card_id, attack_id}", "迎击牌额外效果（移动1格、护甲+2）"),

    # 伤害
    ("伤害", "HOOK_DAMAGE_MODIFIER_WINDOW", "ON_DAMAGE_MODIFIER_WINDOW", "伤害修正窗口触发，可修改本次伤害产生的损伤数量", "AttackService", "{attack_id, target_id, damage_tokens}", "损伤-1等修正效果"),
    ("伤害", "HOOK_DAMAGE_DEALT", "ON_DAMAGE_DEALT", "伤害造成后触发", "GameActions", "{target_id, amount, damage_type}", "伤害减免/伤害反馈效果"),
    ("伤害", "HOOK_OWNER_TAKE_DAMAGE", "ON_OWNER_TAKE_DAMAGE", "我方受到伤害后触发", "EffectEngine", "{mech_id, amount, source}", "受伤后触发效果"),

    # 损伤
    ("损伤", "HOOK_BEFORE_DAMAGE_TOKEN_PLACED", "ON_BEFORE_DAMAGE_TOKEN_PLACED", "损伤放置前触发（可拦截/重定向）", "GameActions", "{target_id, slot_id, cancelled, forced_slot_id}", "损伤重定向/减免"),
    ("损伤", "HOOK_AFTER_DAMAGE_TOKEN_PLACED", "ON_AFTER_DAMAGE_TOKEN_PLACED", "损伤放置后触发", "GameActions", "{target_id, slot_id, amount}", "损伤放置后效果"),

    # 装备
    ("装备", "HOOK_EQUIPMENT_BROKEN", "ON_EQUIPMENT_BROKEN", "装备因损伤超过耐久而破坏时触发", "GameActions", "{mech_id, slot_id, card_id, damage_tokens, durability}", "破坏回复效果"),
    ("装备", "HOOK_EQUIPMENT_DISCARDED_FROM_SLOT", "ON_EQUIPMENT_DISCARDED_FROM_SLOT", "装备从区域弃置时触发", "GameActions", "{card_id, slot_id, reason}", "弃置时清除损伤/抽牌等"),
    ("装备", "HOOK_EQUIPMENT_BROKEN_BY_DAMAGE", "ON_EQUIPMENT_BROKEN_BY_DAMAGE", "装备因损伤弃置时触发（比EQUIPMENT_BROKEN更具体）", "GameActions", "{card_id, slot_id, damage_tokens}", "因损伤弃置时的特殊效果"),

    # 胜负
    ("胜负", "HOOK_MECH_DESTROYED", "ON_MECH_DESTROYED", "机甲被摧毁时触发", "GameActions", "{mech_id, owner_player_id, source}", "终局检查"),

    # 事件
    ("事件", "HOOK_EVENT_TIMER_TICK", "ON_EVENT_TIMER_TICK", "事件计时器递减时触发", "EventTimerService", "{event_card_id, timer}", "计时器递减效果"),
    ("事件", "HOOK_EVENT_TIMER_ZERO", "ON_EVENT_TIMER_ZERO", "事件计时器归零时触发", "EventTimerService", "{event_card_id, mech_id}", "计时器到期效果（如弃置事件牌）"),

    # 地图
    ("地图", "HOOK_MECH_MOVED", "ON_MECH_MOVED", "机甲移动后触发", "MapService", "{mech_id, from, to, distance}", "移动后效果（移动8格回复动力）"),
    ("地图", "HOOK_MECH_LEAVING_CELL", "ON_MECH_LEAVING_CELL", "机甲离开格子时触发", "MapService", "{mech_id, leaving_cell_pos}", "离开设陷阱"),
    ("地图", "HOOK_MAP_GOLD_MARKER_RESOLVED", "ON_MAP_GOLD_MARKER_RESOLVED", "金币标记结算后触发", "MarkerService", "{player_id, amount, roll_result}", "金币获取效果"),
    ("地图", "HOOK_MAP_EVENT_MARKER_RESOLVED", "ON_MAP_EVENT_MARKER_RESOLVED", "事件标记结算后触发", "MarkerService", "{player_id, event_card_id}", "事件牌触发"),
    ("地图", "HOOK_TRAP_EXPLODED", "ON_TRAP_EXPLODED", "陷阱爆炸时触发", "MarkerService", "{cell_id, damage, range}", "陷阱效果"),
    ("地图", "HOOK_MAP_MARKERS_RESET", "ON_MAP_MARKERS_RESET", "地图标记重置时触发", "MarkerService", "{}", "标记重置效果"),

    # 商店
    ("商店", "HOOK_SHOP_INITIALIZED", "ON_SHOP_INITIALIZED", "商店初始化时触发", "ShopService", "{shop_slots}", "商店初始化效果"),
    ("商店", "HOOK_SHOP_CARD_BOUGHT", "ON_SHOP_CARD_BOUGHT", "商店购买卡牌后触发", "ShopService", "{player_id, card_id, cost}", "购买后效果"),
    ("商店", "HOOK_SHOP_REFRESHED", "ON_SHOP_REFRESHED", "商店刷新后触发", "ShopService", "{shop_slots}", "刷新效果"),
    ("商店", "HOOK_SHOP_REFILLED", "ON_SHOP_REFILLED", "商店补货后触发", "ShopService", "{shop_slots}", "补货效果"),

    # 扩展
    ("扩展", "HOOK_STAT_RECALCULATE", "ON_STAT_RECALCULATE", "属性重算时触发（STATIC效果使用）", "EffectRegistry", "{mech_id, slot_id}", "持续属性修正（护甲+2、动力+1等）"),
    ("扩展", "HOOK_STATUS_CHECK", "ON_STATUS_CHECK", "状态检查时触发", "GameState", "{target_id, status_type}", "状态查询/验证"),
    ("扩展", "HOOK_SHOP_ACTION", "ON_SHOP_ACTION", "商店行动时触发", "ShopService", "{action_type, player_id}", "商店行动效果"),
    ("扩展", "HOOK_EQUIPMENT_SOLD", "ON_EQUIPMENT_SOLD", "装备被卖出时触发", "CardSetService", "{player_id, card_id, price}", "卖出效果"),
    ("扩展", "HOOK_ACTION_CARD_DRAWN", "ON_ACTION_CARD_DRAWN", "行动牌被抽取时触发", "GameActions", "{player_id, card_id}", "抽牌触发效果"),
    ("扩展", "HOOK_ENERGY_APPLIED_TO_WEAPON", "ON_ENERGY_APPLIED_TO_WEAPON", "聚能效果应用到武器时触发", "GameActions", "{mech_id, weapon_id, delta}", "聚能额外效果（范围+1/威力+3）"),
    ("扩展", "HOOK_BEFORE_HEAL", "ON_BEFORE_HEAL", "即将回复生命时触发（拦截点）", "GameActions", "{mech_id, amount}", "回复拦截/转化"),
    ("扩展", "HOOK_BEFORE_REMOVE_DAMAGE_TOKENS", "ON_BEFORE_REMOVE_DAMAGE_TOKENS", "即将移除损伤时触发（拦截点）", "GameActions", "{mech_id, slot_id, amount}", "损伤移除拦截/转化"),
    ("扩展", "HOOK_OTHER_MECH_GAIN_GOLD", "ON_OTHER_MECH_GAIN_GOLD", "其他机甲获得金币时触发", "GameActions", "{other_player_id, amount}", "监视对手金币"),
    ("扩展", "HOOK_OTHER_MECH_DRAW_ACTION", "ON_OTHER_MECH_DRAW_ACTION", "其他机甲抽行动牌时触发", "GameActions", "{other_player_id, card_id}", "监视对手抽牌"),
]

row = 2
current_category = ""
for data in hooks_data:
    cat = data[0]
    if cat != current_category:
        current_category = cat
        # Insert category row
        ws1.cell(row=row, column=1, value=cat)
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers1))
        style_row(ws1, row, len(headers1), font=category_font, fill=category_fill)
        row += 1

    for i, val in enumerate(data, 1):
        ws1.cell(row=row, column=i, value=val if i != 1 else "")
    style_row(ws1, row, len(headers1))
    row += 1

# Column widths
ws1.column_dimensions["A"].width = 14
ws1.column_dimensions["B"].width = 40
ws1.column_dimensions["C"].width = 38
ws1.column_dimensions["D"].width = 42
ws1.column_dimensions["E"].width = 22
ws1.column_dimensions["F"].width = 40
ws1.column_dimensions["G"].width = 36

# ═══════════════════════════════════════════
# Sheet 2: 原子动作(Atomic Action)一览表
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("原子动作(AtomicAction)")

headers2 = ["分类", "动作类型", "参数", "执行逻辑", "触发的结果Hook", "备注"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, len(headers2))

actions_data = [
    # 攻击相关
    ("攻击相关", "START_ATTACK_DECLARE_ATTACK", "attacker_id, target_id, weapon_id, attack_card_id", "委托AttackService执行攻击声明，创建攻击上下文", "→ ON_ATTACK_DECLARED", "核心攻击启动动作"),
    ("攻击相关", "MODIFY_ATTACK_POWER", "attack_id, delta, duration", "修改当前攻击的威力值（可正可负），记录modifier", "→ (内部状态修改)", "用于威力+4/-5等修正"),
    ("攻击相关", "MODIFY_ATTACK_RANGE", "attack_id, delta, duration, min", "修改当前攻击的范围值，记录modifier", "→ (内部状态修改)", "用于范围+1/-2等修正"),
    ("攻击相关", "NEGATE_ATTACK", "attack_id", "将攻击标记为取消（若非unnegatable），设置cancelled=true, result=negated", "→ ON_ATTACK_NEGATED", "识破/无效攻击"),
    ("攻击相关", "SET_ATTACK_UNNEGATABLE", "attack_id", "将攻击标记为不可否定(unnegatable=true)", "→ (内部状态修改)", "预判牌防止攻击被无效"),
    ("攻击相关", "APPLY_CANNOT_RESPOND", "target_id, attack_id, duration", "对目标施加CANNOT_RESPOND状态，使其无法响应攻击", "→ ON_STATUS_ADDED", "锁定牌使目标不可响应"),
    ("攻击相关", "APPLY_OR_CHECK_LOCKED", "target_id, mode(apply/check), duration", "mode=apply: 施加LOCKED状态；mode=check: 检查是否已被锁定", "→ ON_STATUS_ADDED (apply时)", "锁定/检查锁定效果"),
    ("攻击相关", "OPEN_OR_USE_RESPONSE", "mode(open/use), attack_id, player_id, response_card_id", "open: 触发响应窗口hook；use: 从手牌移除迎击牌并弃置，标记攻击已响应", "→ ON_ATTACK_RESPONSE_WINDOW / ON_CARD_DISCARDED", "响应窗口控制"),
    ("攻击相关", "CONSUME_NEXT_ATTACK_POWER_BUFF", "attack_id, attacker_id, weapon_id", "查找机甲上的NEXT_ATTACK_POWER_BUFF状态，将delta加到攻击威力上，消耗该状态", "→ (调用MODIFY_ATTACK_POWER)", "聚能效果消耗"),

    # 属性修改
    ("属性修改", "MODIFY_ARMOR", "mech_id, delta, slot_id, duration", "在机甲statuses中添加ARMOR_MODIFIER状态", "→ ON_STATUS_ADDED", "护甲+5/-2等"),
    ("属性修改", "MODIFY_MECH_POWER", "mech_id, delta, duration", "立即修改机甲动力值(clamp到0~max)，如有duration则添加POWER_MODIFIER状态", "→ ON_POWER_CHANGED / ON_STATUS_ADDED", "动力+5/+3等"),
    ("属性修改", "SPEND_POWER", "mech_id, amount, reason", "支付指定动力值，不足则返回false", "→ ON_POWER_SPENT / ON_POWER_CHANGED", "动力消耗"),
    ("属性修改", "RESTORE_POWER", "mech_id, amount/full, reason", "恢复动力值（full=恢复到上限）", "→ ON_POWER_RESTORED / ON_POWER_CHANGED", "回复动力效果"),
    ("属性修改", "RESTORE_WEAPON_POWER", "weapon_id, amount/full, max_value", "恢复武器威力到耐久上限或指定值", "→ ON_WEAPON_POWER_RESTORED", "武器威力回复"),

    # 抽牌/获得
    ("抽牌/获得", "DRAW_ACTION", "player_id, count, reason", "从行动牌堆抽指定数量行动牌，逐张触发hook", "→ ON_CARD_DRAWN / ON_ACTION_CARD_DRAWN / ON_DRAW_FINISHED", "抽行动牌"),
    ("抽牌/获得", "DRAW_EQUIPMENT", "player_id, count, deck_type, reason", "从装备牌堆抽指定数量装备牌", "→ ON_CARD_DRAWN / ON_EQUIPMENT_CARD_DRAWN / ON_DRAW_FINISHED", "抽装备牌"),
    ("抽牌/获得", "GAIN_SPECIFIC_CARD", "player_id, card_def_id, zone", "创建指定card_def的实例并放入玩家手牌", "→ ON_CARD_GAINED", "获得指定卡牌"),
    ("抽牌/获得", "RANDOM_DRAW_FROM_DISCARD_OR_DECK", "player_id, count, source_zone, card_kind", "从弃牌堆或指定牌堆随机抽取", "→ ON_CARD_GAINED", "回收/回忆效果"),
    ("抽牌/获得", "TRANSFER_ACTION_CARDS", "from_player_id, to_player_id, card_ids, count", "将行动牌从一个玩家转移给另一个玩家", "→ ON_CARD_TRANSFERRED", "偷取行动牌"),
    ("抽牌/获得", "GAIN_GOLD", "player_id, amount, reason", "玩家获得指定数量金币", "→ ON_GOLD_GAINED / ON_GOLD_CHANGED", "获得金币"),
    ("抽牌/获得", "SPEND_GOLD", "player_id, amount, reason", "支付指定数量金币，不足返回false", "→ ON_GOLD_SPENT / ON_GOLD_CHANGED", "支付金币"),
    ("抽牌/获得", "SHOP_BUY_MODIFIER", "player_id, delta, multiplier, scope, duration", "添加商店购买修正（折扣/溢价）", "→ ON_STATUS_ADDED", "折扣效果"),

    # 伤害/损伤
    ("伤害/损伤", "DEAL_DAMAGE", "target_id, amount, damage_type", "直接减少机甲HP，HP≤0则摧毁", "→ ON_DAMAGE_DEALT / ON_MECH_DESTROYED", "直接伤害效果"),
    ("伤害/损伤", "PLACE_DAMAGE_TOKENS", "target_id, amount, chooser_player_id, prefer_part_slot, slot_id", "按优先级放置损伤标记，逐枚触发before/after hook", "→ ON_BEFORE_DAMAGE_TOKEN_PLACED / ON_AFTER_DAMAGE_TOKEN_PLACED", "放置损伤标记"),
    ("伤害/损伤", "MODIFY_DAMAGE_TOKENS", "damage_context_id, delta", "修改当前伤害上下文中的损伤数量", "→ (内部状态修改)", "损伤-1/-2等修正"),
    ("伤害/损伤", "REMOVE_DAMAGE_TOKENS", "mech_id, slot_id, amount, all_in_slot, from_other_slots", "从指定区域移除损伤标记（先移区域损伤再移装备损伤）", "→ ON_DAMAGE_TOKEN_REMOVED", "维修/清除损伤"),
    ("伤害/损伤", "REDIRECT_DAMAGE_TOKENS", "damage_context_id, from_slot_id, to_slot_id, amount", "重定向损伤标记到其他区域", "→ (内部状态修改)", "损伤重定向效果"),
    ("伤害/损伤", "HEAL_HP", "mech_id, amount", "恢复机甲HP（不超过max_hp）", "→ ON_HP_HEALED", "回复生命"),

    # 移动/设置
    ("移动/设置", "MOVE_MECH", "mech_id, target_cell_id, ignore_cost, use_current_power, cells, adjacent", "委托MapService移动机甲到指定格子", "→ ON_MECH_MOVED", "移动效果"),
    ("移动/设置", "SET_CARD_TO_SLOT", "card_id, mech_id, slot_id, face_down", "将卡牌设置到指定槽位，刷新效果注册", "→ ON_EQUIPMENT_SET / ON_EVENT_SET", "装备/事件设置"),
    ("移动/设置", "PLACE_OR_TRIGGER_TRAP", "mode(place/trigger), cell_id, marker_id", "放置陷阱标记到地图格或触发已有陷阱", "→ ON_MAP_MARKER_PLACED / (触发陷阱)", "陷阱放置/触发"),

    # 弃牌/破坏
    ("弃牌/破坏", "DISCARD_CARD", "card_id, reason", "注销卡牌效果→从所有区域移除→加入弃牌堆", "→ ON_CARD_DISCARDED", "弃牌核心动作"),
    ("弃牌/破坏", "DISCARD_ACTION_CARD", "player_id, card_id, count", "从玩家行动手牌中弃置指定数量行动牌", "→ (调用DISCARD_CARD)", "弃置行动牌"),
    ("弃牌/破坏", "DESTROY_CARD", "card_id, reason", "先触发ON_CARD_DESTROYED hook，再弃置卡牌", "→ ON_CARD_DESTROYED / ON_CARD_DISCARDED", "破坏卡牌"),
    ("弃牌/破坏", "PLAY_AS_CARD", "player_id, original_card_id, virtual_card_id, targets", "弃置原牌，创建虚拟牌实例并执行其效果", "→ ON_CARD_PLAYED", "作为其他牌打出"),

    # 状态
    ("状态", "ADD_STATUS", "target_id, status", "向目标添加状态", "→ ON_STATUS_ADDED", "添加任意状态"),
    ("状态", "REMOVE_STATUS", "target_id, status_id, status_type", "从目标移除指定状态", "→ ON_STATUS_REMOVED", "移除状态"),
    ("状态", "ADD_RULE_MODIFIER", "rule, rule_id, rule_type, value, duration", "添加规则修正到game_state.rule_modifiers", "→ ON_RULE_MODIFIER_ADDED", "规则修正（陷阱次数、可卖出等）"),

    # 事件/计时
    ("事件/计时", "REDUCE_EVENT_TIMER", "event_card_id, amount", "减少事件卡牌的计时器值，归零时触发ON_EVENT_TIMER_ZERO", "→ ON_EVENT_TIMER_TICK / ON_EVENT_TIMER_ZERO", "事件计时递减"),
    ("事件/计时", "SET_EVENT_TIMER", "event_card_id, value", "直接设置事件卡牌计时器值", "→ ON_EVENT_TIMER_SET", "设置事件计时"),
    ("事件/计时", "TRACK_EVENT_PROGRESS", "event_card_id, metric, delta", "追踪事件进度（计数器+delta）", "→ ON_EVENT_PROGRESS_CHANGED", "事件进度追踪"),

    # 其他
    ("其他", "APPLY_ENERGY_TO_WEAPON", "mech_id, weapon_id, delta", "对武器施加NEXT_ATTACK_POWER_BUFF状态（下次攻击威力+delta，一次性消耗）", "→ ON_ENERGY_APPLIED_TO_WEAPON", "聚能效果"),
    ("其他", "STEAL_ACTION_CARD", "from_player_id, to_player_id, count", "从对手手牌偷取行动牌", "→ ON_CARD_TRANSFERRED", "偷取行动牌"),
    ("其他", "PLACE_TRAP_MARKER", "cell_id, mech_id, damage, range, tokens", "在地图格上放置TRAP类型标记", "→ ON_MAP_MARKER_PLACED", "放置陷阱标记"),
    ("其他", "CONVERT_WEAPON_KIND", "weapon_id, new_kind", "转换武器类型（如远程→近战）", "→ ON_STATUS_ADDED", "武器类型转换"),
    ("其他", "REVEAL_OR_PEEK_CARD", "player_id, mode(reveal/peek), card_ids", "揭示或窥视卡牌", "→ ON_CARD_PEEKED / ON_CARD_REVEALED", "卡牌信息揭示"),
    ("其他", "ROLL_D6", "store_key", "掷D6骰子(1-6)，结果可存入temp_values", "→ ON_DICE_ROLLED", "随机效果"),
    ("其他", "TOGGLE_AURA_TARGET", "target_id, aura_id, enabled", "启用/禁用光环目标", "→ ON_AURA_TARGET_CHANGED", "光环效果控制"),
    ("其他", "CUSTOM_EFFECT_CHECK_TEXT", "effect_id, text", "自定义效果文本检查（兜底）", "→ ON_CUSTOM_EFFECT_REQUIRED", "自定义效果"),

    # 阶段1新增
    ("阶段1扩展", "PLACE_DAMAGE_TOKENS_ON_SLOT", "mech_id, slot_id, amount", "直接在指定区域/此牌上放置损伤（不经过优先级选择）", "→ ON_AFTER_DAMAGE_TOKEN_PLACED", "自损效果"),
    ("阶段1扩展", "PLAY_CARD_AS_TYPE", "player_id, card_id, as_type", "将行动牌当作指定类型使用", "→ ON_CARD_PLAYED", "牌型转换使用"),
    ("阶段1扩展", "MODIFY_ACTION_HAND_LIMIT", "player_id, delta, duration", "修改行动牌手牌上限", "→ ON_STATUS_ADDED", "手牌上限调整"),
    ("阶段1扩展", "MODIFY_ATTACK_COUNT", "mech_id, delta, duration", "修改本回合可攻击次数", "→ ON_STATUS_ADDED", "攻击次数调整"),
    ("阶段1扩展", "INCREMENT_VARIABLE", "variable_name, delta, player_id, mech_id", "自定义计数器+delta", "→ (内部状态修改)", "计数器累加"),
    ("阶段1扩展", "CHOOSE_ONE", "chosen_effect_id", "效果路由：根据玩家选择执行其中一个效果", "→ (委托EffectEngine)", "选择效果分支"),
    ("阶段1扩展", "FORCE_MECH_ACTION", "target_mech_id, action_type", "强制其他机甲执行行动", "→ (添加forced_action状态)", "联合效果"),
    ("阶段1扩展", "TREAT_CARD_AS_NAMED_TYPE", "player_id, card_id, named_type", "将牌视作指定命名类型使用", "→ ON_CARD_PLAYED", "牌型视为"),
    ("阶段1扩展", "GRANT_EFFECT_TO_FACTION", "faction, effect_id", "使指定阵营所有机甲获得效果", "→ (添加faction_effect_grant状态)", "阵营光环"),
    ("阶段1扩展", "TOGGLE_EFFECT_ON_MECH", "mech_id, effect_id, toggle(cancel/restore)", "取消/恢复机甲上的效果", "→ (修改状态)", "效果开关"),
    ("阶段1扩展", "NEGATE_EQUIPMENT_EFFECT", "target_card_id, duration", "使装备效果无效（设置disabled=true）", "→ (修改卡牌状态)", "装备无效化"),
    ("阶段1扩展", "MOVE_WITHOUT_POWER", "mech_id, cells", "无视动力消耗移动指定格数", "→ (委托MapService)", "免费移动"),
    ("阶段1扩展", "MODIFY_WEAPON_POWER", "weapon_id, delta, duration", "修改武器威力（非仅回复）", "→ (修改might_modifiers)", "武器威力修正"),
    ("阶段1扩展", "SET_WEAPON_STATS", "weapon_id, might, range", "直接设置武器属性为指定值", "→ (修改weapon属性)", "武器属性覆盖"),
    ("阶段1扩展", "CONVERT_ARMOR_TO_POWER", "mech_id, armor_amount, draw_per_2_armor", "将护甲转化为动力，每2点护甲抽1张行动牌", "→ (调用modify_mech_power/draw_action_cards)", "护甲转动力"),
    ("阶段1扩展", "REDIRECT_HEAL_TO_DAMAGE", "target_mech_id, amount", "将回复生命改为受到等量伤害", "→ ON_DAMAGE_DEALT", "回复转伤害"),
    ("阶段1扩展", "REDIRECT_REMOVE_TO_PLACE_TOKENS", "target_mech_id, amount, slot_id", "将移除损伤改为放置损伤", "→ (调用place_damage_tokens)", "移除转放置"),
    ("阶段1扩展", "MODIFY_NEXT_DAMAGE_DEALT", "mech_id, delta", "设置下次造成的伤害+N", "→ (存入variables)", "下次伤害加成"),
    ("阶段1扩展", "ADD_WEAPON_TAG", "weapon_id, tag, duration", "给武器添加名称标签（热能/光束）", "→ (修改weapon tags)", "武器标签"),
    ("阶段1扩展", "DECLARE_CARD_TYPE", "player_id, declared_type", "宣言行动牌类型", "→ (添加declared_card_type状态)", "牌型宣言"),
    ("阶段1扩展", "DRAW_ADVANCED_EQUIPMENT", "player_id, count", "从高级装备牌堆抽牌", "→ (委托DeckService)", "抽高级装备"),
    ("阶段1扩展", "PLACE_CARD_IN_DECK_FACE_UP", "player_id, card_ids, top_card_id", "将牌正面朝上放入行动牌堆", "→ (修改牌堆)", "正面朝上入堆"),

    # 机师效果
    ("机师效果", "SWAP_HAND_LIMIT_AND_ATTACK_COUNT", "player_id, mech_id", "互换行动牌上限与回合攻击数，交换后立即抽牌到新上限", "→ ON_STATUS_ADDED", "机师特殊能力"),
]

row = 2
current_category = ""
for data in actions_data:
    cat = data[0]
    if cat != current_category:
        current_category = cat
        ws2.cell(row=row, column=1, value=cat)
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers2))
        style_row(ws2, row, len(headers2), font=category_font, fill=category_fill)
        row += 1

    for i, val in enumerate(data, 1):
        ws2.cell(row=row, column=i, value=val if i != 1 else "")
    style_row(ws2, row, len(headers2))
    row += 1

ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 48
ws2.column_dimensions["D"].width = 55
ws2.column_dimensions["E"].width = 45
ws2.column_dimensions["F"].width = 28

# ═══════════════════════════════════════════
# Sheet 3: 行动牌效果执行逻辑
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("行动牌效果逻辑")

headers3 = [
    "卡牌ID", "卡牌名称", "类型", "稀有度", "数量",
    "效果发动时点", "效果ID", "效果模式",
    "触发条件", "目标规则", "费用",
    "执行的动作", "执行逻辑说明", "效果文本"
]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, len(headers3))

action_cards = [
    {
        "id": "action_001_进攻", "name": "进攻", "type": "攻击", "rarity": "N", "count": 14,
        "effects": [
            {
                "hook": "ON_ATTACK_CARD_PLAYED（攻击牌打出时）",
                "effect_id": "basic_attack_single",
                "mode": "PASSIVE（被动，自动触发）",
                "conditions": "ALWAYS（无条件）",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "START_ATTACK_DECLARE_ATTACK",
                "logic": "当玩家打出进攻牌时，自动触发攻击声明流程：选择1把武器→选择1台范围内的敌方机甲→创建攻击上下文→进入响应窗口",
            },
        ],
        "text": "选择1把武器对1台范围内的机甲发动攻击。"
    },
    {
        "id": "action_002_强袭", "name": "强袭", "type": "攻击", "rarity": "R", "count": 8,
        "effects": [
            {
                "hook": "ON_ATTACK_CARD_PLAYED",
                "effect_id": "basic_attack_single",
                "mode": "PASSIVE",
                "conditions": "ALWAYS",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "START_ATTACK_DECLARE_ATTACK",
                "logic": "同进攻：发起攻击声明",
            },
            {
                "hook": "ON_ATTACK_RESPONSE_WINDOW（攻击响应窗口）",
                "effect_id": "move_current_power_after_response_before_resolution",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_ATTACKER（效果来源的拥有者是攻击方）",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "MOVE_MECH(use_current_power=true)",
                "logic": "在防守方响应后、攻击结算前，攻击方可用当前动力进行移动，之后再结算攻击距离和伤害",
            },
        ],
        "text": "选择1把武器对1台范围内的机甲发动攻击，并可以在目标响应后用当前的动力进行移动，之后再结算本次攻击。"
    },
    {
        "id": "action_003_猛击", "name": "猛击", "type": "攻击", "rarity": "R", "count": 8,
        "effects": [
            {
                "hook": "ON_ATTACK_CARD_PLAYED",
                "effect_id": "basic_attack_single",
                "mode": "PASSIVE",
                "conditions": "ALWAYS",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "START_ATTACK_DECLARE_ATTACK",
                "logic": "发起攻击声明",
            },
            {
                "hook": "ON_ATTACK_MODIFIER_WINDOW（攻击修正窗口）",
                "effect_id": "attack_power_plus_4_this_attack",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_ATTACKER",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "MODIFY_ATTACK_POWER(delta=4, duration=THIS_ATTACK)",
                "logic": "在攻击修正窗口中，将本次攻击威力+4（仅本次攻击有效）",
            },
        ],
        "text": "选择1把武器对1台范围内的机甲发动攻击，使本次攻击威力+4。"
    },
    {
        "id": "action_004_破甲", "name": "破甲", "type": "攻击", "rarity": "R", "count": 8,
        "effects": [
            {
                "hook": "ON_ATTACK_CARD_PLAYED",
                "effect_id": "basic_attack_single",
                "mode": "PASSIVE",
                "conditions": "ALWAYS",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "START_ATTACK_DECLARE_ATTACK",
                "logic": "发起攻击声明",
            },
            {
                "hook": "ON_ATTACK_HIT（攻击命中时）",
                "effect_id": "on_hit_add_2_damage_markers",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_ATTACKER + PAYLOAD_ATTACK_HIT",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "PLACE_DAMAGE_TOKENS(amount=2, target=$payload.target_id)",
                "logic": "若攻击命中，对目标额外放置2枚损伤标记",
            },
        ],
        "text": "选择1把武器对1台范围内的机甲发动攻击，若本次攻击命中则可额外设置2枚损伤。"
    },
    {
        "id": "action_005_双连", "name": "双连", "type": "攻击", "rarity": "R", "count": 6,
        "effects": [
            {
                "hook": "ON_ATTACK_CARD_PLAYED",
                "effect_id": "attack_one_weapon_one_or_two_targets",
                "mode": "PASSIVE",
                "conditions": "ALWAYS",
                "targets": "CHOOSE_ENEMY_MECH",
                "cost": "无",
                "actions": "START_ATTACK_DECLARE_ATTACK(max_targets=2)",
                "logic": "与进攻不同：可选择1~2台范围内的敌方机甲发动攻击（同一武器对多目标）",
            },
        ],
        "text": "选择1把武器对1~2台范围内的机甲发动攻击。"
    },
    {
        "id": "action_006_闪击", "name": "闪击", "type": "攻击", "rarity": "SR", "count": 6,
        "effects": [
            {
                "hook": "ON_ATTACK_CARD_PLAYED",
                "effect_id": "basic_attack_single",
                "mode": "PASSIVE",
                "conditions": "ALWAYS",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "START_ATTACK_DECLARE_ATTACK",
                "logic": "发起攻击声明",
            },
            {
                "hook": "ON_ATTACK_RESOLVED（攻击结算后）",
                "effect_id": "discard_action_repeat_same_attack",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_ATTACKER",
                "targets": "NO_TARGET",
                "cost": "DISCARD_ACTION_CARD(count=1)",
                "actions": "START_ATTACK_DECLARE_ATTACK(repeat_last_attack=true)",
                "logic": "攻击结算后，可弃置1张行动牌作为费用，然后用相同武器对相同目标再次发动攻击",
            },
        ],
        "text": "选择1把武器对1台范围内的机甲发动攻击；该攻击结算后，可以弃置1张行动牌，选择相同的武器对相同的目标再次发动攻击。"
    },
    {
        "id": "action_007_预判", "name": "预判", "type": "攻击", "rarity": "SSR", "count": 2,
        "effects": [
            {
                "hook": "ON_ATTACK_CARD_PLAYED",
                "effect_id": "basic_attack_single",
                "mode": "PASSIVE",
                "conditions": "ALWAYS",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "START_ATTACK_DECLARE_ATTACK",
                "logic": "发起攻击声明",
            },
            {
                "hook": "ON_ATTACK_HIT",
                "effect_id": "apply_lock_effect_to_target",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_ATTACKER + PAYLOAD_ATTACK_HIT",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "APPLY_OR_CHECK_LOCKED(apply=true, target=$payload.target_id)",
                "logic": "攻击命中后，对目标施加LOCKED锁定效果",
            },
            {
                "hook": "ON_ATTACK_HIT",
                "effect_id": "discard_target_action_card_1",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_ATTACKER + PAYLOAD_ATTACK_HIT",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "STEAL_ACTION_CARD(from_target=true, count=1, discard=true)",
                "logic": "攻击命中后，弃置目标1张行动牌",
            },
            {
                "hook": "ON_ATTACK_CARD_PLAYED",
                "effect_id": "attack_cannot_be_nullified",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_ATTACKER",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "SET_ATTACK_UNNEGATABLE",
                "logic": "此牌发动的攻击被标记为不可否定，无法被识破等效果无效",
            },
        ],
        "text": "选择1把武器对1台范围内的机甲发动攻击，对目标施加锁定效果，并可以弃置目标1张行动牌。此牌发动的攻击不会被无效。"
    },
    {
        "id": "action_008_回避", "name": "回避", "type": "迎击", "rarity": "N", "count": 12,
        "effects": [
            {
                "hook": "ON_ATTACK_RESPONSE_WINDOW",
                "effect_id": "evade_half_power",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_TARGET（效果来源的拥有者是攻击目标）",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "MOVE_MECH(power_fraction=0.5)",
                "logic": "响应窗口中，被攻击方可用当前1/2的动力（向下取整）进行移动，以脱离攻击范围",
            },
        ],
        "text": "响应对我方的攻击，可以用当前1/2的动力（向下取整）进行移动。"
    },
    {
        "id": "action_009_防御", "name": "防御", "type": "迎击", "rarity": "N", "count": 8,
        "effects": [
            {
                "hook": "ON_ATTACK_RESPONSE_WINDOW",
                "effect_id": "defend_armor_bonus_5",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_TARGET",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "MODIFY_ARMOR(delta=5, duration=THIS_ATTACK)",
                "logic": "响应窗口中，被攻击方护甲+5，持续到本次攻击结算结束",
            },
            {
                "hook": "ON_DAMAGE_MODIFIER_WINDOW（伤害修正窗口）",
                "effect_id": "reduce_attack_damage_marker_1",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_TARGET",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "MODIFY_DAMAGE_TOKENS(delta=-1)",
                "logic": "伤害修正窗口中，减少本次攻击产生的1损伤",
            },
        ],
        "text": "响应对我方的攻击，在本次攻击结算前使机甲护甲+5，减少该攻击产生的1损伤。"
    },
    {
        "id": "action_010_反击", "name": "反击", "type": "迎击", "rarity": "R", "count": 6,
        "effects": [
            {
                "hook": "ON_ATTACK_RESPONSE_WINDOW",
                "effect_id": "evade_half_power",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_TARGET",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "MOVE_MECH(power_fraction=0.5)",
                "logic": "先用1/2动力移动",
            },
            {
                "hook": "ON_ATTACK_RESOLVED",
                "effect_id": "counterattack_after_resolution",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_TARGET",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "START_ATTACK_DECLARE_ATTACK",
                "logic": "攻击结算后，被攻击方可选择1把武器对1台攻击范围内的机甲发动反击",
            },
        ],
        "text": "响应攻击，可以用当前1/2的动力（向下取整）进行移动；该攻击结算后，可以选择1把武器对1台攻击范围内的机甲发动攻击。"
    },
    {
        "id": "action_011_疾行", "name": "疾行", "type": "迎击", "rarity": "R", "count": 6,
        "effects": [
            {
                "hook": "ON_ATTACK_RESPONSE_WINDOW",
                "effect_id": "evade_full_power",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_TARGET",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "MOVE_MECH(use_current_power=true)",
                "logic": "响应窗口中，被攻击方可用当前全部动力进行移动（比回避更灵活）",
            },
        ],
        "text": "响应对我方的攻击，可以用当前的动力进行移动。"
    },
    {
        "id": "action_012_识破", "name": "识破", "type": "迎击", "rarity": "SSR", "count": 2,
        "effects": [
            {
                "hook": "ON_ATTACK_RESPONSE_WINDOW",
                "effect_id": "nullify_attack",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_TARGET",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "NEGATE_ATTACK",
                "logic": "直接将攻击标记为无效（cancelled=true）",
            },
            {
                "hook": "ON_ATTACK_RESPONSE_WINDOW",
                "effect_id": "gain_attacker_action_card_1",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_TARGET",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "STEAL_ACTION_CARD(from_attacker=true, count=1)",
                "logic": "偷取攻击方1张行动牌加入自己手牌",
            },
            {
                "hook": "ON_ATTACK_RESPONSE_WINDOW",
                "effect_id": "evade_full_power",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_TARGET",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "MOVE_MECH(use_current_power=true)",
                "logic": "用当前全部动力移动",
            },
            {
                "hook": "ON_ATTACK_CARD_PLAYED",
                "effect_id": "ignore_lock_when_played",
                "mode": "PASSIVE",
                "conditions": "ALWAYS",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "APPLY_OR_CHECK_LOCKED(ignore_lock=true)",
                "logic": "打出此牌时不受锁定状态影响",
            },
        ],
        "text": "响应对我方的攻击，直接无效该攻击并获得攻击方的1张行动牌，之后可以用当前的动力进行移动。打出此牌不受锁定影响。"
    },
    {
        "id": "action_013_维修", "name": "维修", "type": "辅助", "rarity": "N", "count": 12,
        "effects": [
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "repair_heal_life_2",
                "mode": "PASSIVE",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "HEAL_HP(amount=2)",
                "logic": "在拥有者主阶段回复2点HP",
            },
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "repair_remove_damage_2",
                "mode": "PASSIVE",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "REMOVE_DAMAGE_TOKENS(amount=2)",
                "logic": "在拥有者主阶段移除2枚损伤",
            },
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "can_target_adjacent_mecha",
                "mode": "PASSIVE",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "TARGET_IS_ADJACENT",
                "cost": "无",
                "actions": "（无动作，仅扩展目标规则）",
                "logic": "修改目标规则：也可对1格范围内的其他机甲使用",
            },
        ],
        "text": "回复机甲2点生命或移除2枚损伤（也可对1格范围内的其他机甲使用)。"
    },
    {
        "id": "action_014_聚能", "name": "聚能", "type": "辅助", "rarity": "N", "count": 8,
        "effects": [
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "next_attack_power_plus_4_selected_weapon",
                "mode": "PASSIVE",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "CHOOSE_OWN_WEAPON",
                "cost": "无",
                "actions": "CONSUME_NEXT_ATTACK_POWER_BUFF(delta=4)",
                "logic": "在拥有者主阶段，选择我方1把武器，为其添加NEXT_ATTACK_POWER_BUFF状态(+4)。该武器下次攻击时消耗此状态，攻击威力+4",
            },
        ],
        "text": "本回合内选择我方1把武器使其下次发动的攻击威力+4。"
    },
    {
        "id": "action_015_推进", "name": "推进", "type": "辅助", "rarity": "N", "count": 6,
        "effects": [
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "gain_power_5_this_turn",
                "mode": "PASSIVE",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "MODIFY_MECH_POWER(delta=5, duration=THIS_TURN)",
                "logic": "在拥有者主阶段，本回合动力+5（回合结束时恢复）",
            },
            {
                "hook": "ON_ATTACK_CARD_PLAYED",
                "effect_id": "can_play_with_reaction_card",
                "mode": "PASSIVE",
                "conditions": "ALWAYS",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "ADD_RULE_MODIFIER(rule=play_with_reaction, duration=THIS_TURN)",
                "logic": "添加规则修正：此牌也可以与迎击牌一同打出",
            },
        ],
        "text": "本回合使机甲动力+5；此牌也可以与迎击牌一同打出。"
    },
    {
        "id": "action_016_掩护", "name": "掩护", "type": "辅助", "rarity": "N", "count": 5,
        "effects": [
            {
                "hook": "ON_ATTACK_DECLARED（攻击声明时）",
                "effect_id": "cover_reduce_attack_power_5",
                "mode": "PASSIVE",
                "conditions": "ALLY_IN_WEAPON_RANGE_IS_TARGET（武器范围内有友方被指定为目标）",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "MODIFY_ATTACK_POWER(delta=-5, duration=THIS_ATTACK)",
                "logic": "当武器范围内存在机甲（含我方）被攻击时可以打出，使该攻击威力-5",
            },
        ],
        "text": "已设置武器的范围内存在机甲(包括我方)被攻击时可以打出，使该攻击威力-5。"
    },
    {
        "id": "action_017_设陷", "name": "设陷", "type": "辅助", "rarity": "N", "count": 5,
        "effects": [
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "trap_two_chances_this_turn",
                "mode": "PASSIVE",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "ADD_RULE_MODIFIER(rule=trap_chances, value=2, duration=THIS_TURN)",
                "logic": "本回合获得2次陷阱机会",
            },
            {
                "hook": "ON_MECH_LEAVING_CELL（机甲离开格子时）",
                "effect_id": "set_trap_when_leaving_cell",
                "mode": "PASSIVE",
                "conditions": "ALWAYS",
                "targets": "CHOOSE_MAP_CELL_IN_WEAPON_RANGE",
                "cost": "无",
                "actions": "PLACE_TRAP_MARKER(cell_pos=$payload.leaving_cell_pos)",
                "logic": "当任何机甲离开某格子时，可在该格子设置1枚陷阱（消耗1次陷阱机会）",
            },
        ],
        "text": "本回合有2次机会，当机甲离开某格子时可以在该格子上设置1枚陷阱。"
    },
    {
        "id": "action_018_联合", "name": "联合", "type": "辅助", "rarity": "R", "count": 4,
        "effects": [
            {
                "hook": "ON_ATTACK_RESOLVED",
                "effect_id": "allow_other_mecha_attack_after_your_attack",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_ATTACKER",
                "targets": "CHOOSE_ENEMY_MECH_IN_RANGE(range=5)",
                "cost": "无",
                "actions": "FORCE_MECH_ACTION(action_type=attack)",
                "logic": "你攻击结算后，选择其他1台机甲，本回合其也可打出1张攻击牌",
            },
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "discard_self_draw_action_1",
                "mode": "ACTIVE（主动选择使用）",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "NO_TARGET",
                "cost": "DISCARD_ACTION_CARD(count=1)",
                "actions": "DRAW_ACTION(count=1)",
                "logic": "也可以弃置此牌（作为替代用法），然后抽1张行动牌",
            },
        ],
        "text": "选择其他1台机甲，本回合其在你发动攻击结算完成后也可以打出1张攻击牌。也可以弃置此牌，然后抽1张行动牌。"
    },
    {
        "id": "action_019_回收", "name": "回收", "type": "辅助", "rarity": "R", "count": 6,
        "effects": [
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "draw_random_equipment_from_discard",
                "mode": "ACTIVE",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "RANDOM_DRAW_FROM_DISCARD_OR_DECK(type=equipment, count=1)",
                "logic": "主动效果：从装备弃牌堆随机抽1张牌加入手牌",
            },
        ],
        "text": "从装备弃牌堆里随机抽1张牌。"
    },
    {
        "id": "action_020_回忆", "name": "回忆", "type": "辅助", "rarity": "R", "count": 6,
        "effects": [
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "draw_random_action_2_from_discard",
                "mode": "ACTIVE",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "RANDOM_DRAW_FROM_DISCARD_OR_DECK(type=action, count=2)",
                "logic": "主动效果：从行动弃牌堆随机抽2张牌加入手牌",
            },
        ],
        "text": "从行动弃牌堆里随机抽2张牌。"
    },
    {
        "id": "action_021_折扣", "name": "折扣", "type": "辅助", "rarity": "SR", "count": 4,
        "effects": [
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "buy_equipment_at_face_value_twice",
                "mode": "ACTIVE",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "SHOP_BUY_MODIFIER(face_value=true, uses=2)",
                "logic": "添加商店购买修正：本回合2次以原价（面值）购买装备牌的资格",
            },
        ],
        "text": "本回合有2次机会，可以在商店中以原价购买装备牌。"
    },
    {
        "id": "action_022_补给", "name": "补给", "type": "辅助", "rarity": "SR", "count": 5,
        "effects": [
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "draw_action_2_equipment_1",
                "mode": "ACTIVE",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "DRAW_ACTION(count=2) → DRAW_EQUIPMENT(count=1)",
                "logic": "主动效果：先抽2张行动牌，再抽1张装备牌",
            },
        ],
        "text": "抽2张行动牌与1张装备牌。"
    },
    {
        "id": "action_023_锁定", "name": "锁定", "type": "辅助", "rarity": "SR", "count": 4,
        "effects": [
            {
                "hook": "ON_ATTACK_CARD_PLAYED",
                "effect_id": "target_cannot_react_to_your_attacks",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_ATTACKER",
                "targets": "CHOOSE_ENEMY_MECH_IN_RANGE(range=5)",
                "cost": "无",
                "actions": "APPLY_CANNOT_RESPOND(duration=THIS_TURN)",
                "logic": "指定1台敌方机甲，对其施加CANNOT_RESPOND状态，本回合其不能响应你的攻击",
            },
            {
                "hook": "ON_ATTACK_HIT",
                "effect_id": "lock_effect_ends_after_target_hit",
                "mode": "PASSIVE",
                "conditions": "SOURCE_OWNER_IS_ATTACKER + PAYLOAD_ATTACK_HIT",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "REMOVE_STATUS(status_type=cannot_respond, target=$payload.target_id)",
                "logic": "当目标被攻击命中后，移除其CANNOT_RESPOND状态，结束锁定效果",
            },
        ],
        "text": "指定其他1台机甲，本回合其不能响应你发动的攻击(该目标机甲被攻击命中后结束以上效果)。"
    },
    {
        "id": "action_024_觉醒", "name": "觉醒", "type": "辅助", "rarity": "SSR", "count": 1,
        "effects": [
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "gain_prediction_and_insight_from_action_discard",
                "mode": "ACTIVE",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "GAIN_SPECIFIC_CARD(card_ids=[闪击, 识破])",
                "logic": "从行动弃牌堆获得预判与识破各1张",
            },
            {
                "hook": "ON_OWNER_MAIN_PHASE",
                "effect_id": "replace_missing_named_action_and_draw",
                "mode": "ACTIVE",
                "conditions": "IS_OWNER_MAIN_PHASE",
                "targets": "NO_TARGET",
                "cost": "无",
                "actions": "GAIN_SPECIFIC_CARD(fallback_to_choice=true, draw_per_missing=1)",
                "logic": "弃牌堆每缺少1种以上记述的行动牌，可从弃牌堆指定获得1张行动牌，并抽1张行动牌作为补偿",
            },
        ],
        "text": "从行动弃牌堆里获得预判与识破各1张，弃牌堆每缺少以上记述的1种行动牌，则可以从弃牌堆里指定获得1张行动牌，并抽1张行动牌。"
    },
]

row = 2
type_colors = {
    "攻击": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "迎击": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "辅助": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
}

for card in action_cards:
    card_type = card["type"]
    type_fill = type_colors.get(card_type, None)

    for i, eff in enumerate(card["effects"]):
        r = row
        ws3.cell(row=r, column=1, value=card["id"] if i == 0 else "")
        ws3.cell(row=r, column=2, value=card["name"] if i == 0 else "")
        ws3.cell(row=r, column=3, value=card["type"] if i == 0 else "")
        ws3.cell(row=r, column=4, value=card["rarity"] if i == 0 else "")
        ws3.cell(row=r, column=5, value=card["count"] if i == 0 else "")
        ws3.cell(row=r, column=6, value=eff["hook"])
        ws3.cell(row=r, column=7, value=eff["effect_id"])
        ws3.cell(row=r, column=8, value=eff["mode"])
        ws3.cell(row=r, column=9, value=eff["conditions"])
        ws3.cell(row=r, column=10, value=eff["targets"])
        ws3.cell(row=r, column=11, value=eff["cost"])
        ws3.cell(row=r, column=12, value=eff["actions"])
        ws3.cell(row=r, column=13, value=eff["logic"])
        ws3.cell(row=r, column=14, value=card["text"] if i == 0 else "")

        for col in range(1, len(headers3) + 1):
            cell = ws3.cell(row=r, column=col)
            cell.font = body_font
            cell.alignment = wrap_align
            cell.border = thin_border
            if type_fill and i == 0:
                cell.fill = type_fill
        row += 1

ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 10
ws3.column_dimensions["C"].width = 8
ws3.column_dimensions["D"].width = 8
ws3.column_dimensions["E"].width = 6
ws3.column_dimensions["F"].width = 35
ws3.column_dimensions["G"].width = 42
ws3.column_dimensions["H"].width = 24
ws3.column_dimensions["I"].width = 40
ws3.column_dimensions["J"].width = 28
ws3.column_dimensions["K"].width = 24
ws3.column_dimensions["L"].width = 52
ws3.column_dimensions["M"].width = 60
ws3.column_dimensions["N"].width = 55

# ── Save ──
output_path = "f:/Battle-GEAR-S/docs/effect_system_reference.xlsx"
wb.save(output_path)
print(f"Excel saved to: {output_path}")
